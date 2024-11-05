import os
import shutil
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import time

# Directorio de almacenamiento de datos
data_dir = "data"
os.makedirs(data_dir, exist_ok=True)

# Parámetros avanzados por defecto
base_size = 10
max_point_size = 500
min_point_size = 5
label_threshold = 400

# Ruta del archivo plantilla
plantilla = 'config/MeteoData.xlsx'

def obtener_fechas_disponibles_api(latitud, longitud):
    url = f"https://api.open-meteo.com/v1/forecast?latitude={latitud}&longitude={longitud}&daily=temperature_2m_max&timezone=Europe/Madrid"
    intento = 0
    max_reintentos = 5
    while intento < max_reintentos:
        response = requests.get(url)
        if response.status_code == 200:
            data = response.json()
            return data['daily']['time']
        elif response.status_code == 429:
            print("Demasiadas solicitudes. Esperando antes de reintentar...")
            time.sleep(5)
            intento += 1
        else:
            print(f"Error en la solicitud de fechas disponibles: Código {response.status_code}")
            return []
    print("No se pudo obtener las fechas disponibles después de varios intentos.")
    return []


def verificar_archivo_existente(fecha, comunidad):
    fecha_str = fecha.replace("-", "")
    archivo = os.path.join(data_dir, f"MeteoData_{fecha_str}.xlsx")
    
    if os.path.exists(archivo):
        ultima_actualizacion = datetime.fromtimestamp(os.path.getmtime(archivo)).strftime('%Y-%m-%d %H:%M:%S')
        datos_rellenos = verificar_datos_rellenos(archivo, comunidad)
        return True, ultima_actualizacion, datos_rellenos
    return False, None, False

def verificar_datos_rellenos(archivo, comunidad):
    """Verifica si hay datos rellenados en la hoja correspondiente a la comunidad autónoma."""
    wb = load_workbook(archivo)
    if comunidad not in wb.sheetnames:
        return False
    ws = wb[comunidad]
    
    for row in range(2, ws.max_row + 1):  # Asumimos que la fila 1 tiene encabezados
        if ws[f'H{row}'].value is not None:  # Verifica la columna 'H' como referencia
            return True
    return False

def menu_opciones_avanzadas():
    global base_size, max_point_size, min_point_size, label_threshold
    print("\nOpciones avanzadas de parametrización:")
    base_size = int(input(f"Introduce el base_size (actual: {base_size}): ") or base_size)
    max_point_size = int(input(f"Introduce el max_point_size (actual: {max_point_size}): ") or max_point_size)
    min_point_size = int(input(f"Introduce el min_point_size (actual: {min_point_size}): ") or min_point_size)
    label_threshold = int(input(f"Introduce el label_threshold (actual: {label_threshold}): ") or label_threshold)
    print("Parámetros actualizados.\n")

def obtener_datos_meteo(latitud, longitud, fecha):
    url = f"https://api.open-meteo.com/v1/forecast?latitude={latitud}&longitude={longitud}&daily=temperature_2m_max,temperature_2m_min,windspeed_10m_max,windgusts_10m_max,winddirection_10m_dominant,precipitation_sum&timezone=Europe/Madrid"
    response = requests.get(url)
    if response.status_code == 200:
        return response.json()
    else:
        print(f"Error en la solicitud para {fecha}: Código {response.status_code}")
        return None

def copiar_plantilla(fecha):
    fecha_str = fecha.strftime('%Y%m%d')
    archivo_destino = os.path.join(data_dir, f"MeteoData_{fecha_str}.xlsx")
    shutil.copyfile(plantilla, archivo_destino)
    return archivo_destino

def seleccionar_comunidad():
    comunidades = ["ANDALUCIA", "VALENCIA"]
    print("\nComunidades Autónomas disponibles:")
    for i, comunidad in enumerate(comunidades, start=1):
        print(f"{i}. {comunidad}")
    while True:
        try:
            seleccion = int(input("Selecciona la comunidad autónoma (por número): ")) - 1
            if 0 <= seleccion < len(comunidades):
                return comunidades[seleccion]
            else:
                print("Selección no válida. Intenta nuevamente.")
        except ValueError:
            print("Entrada no válida. Por favor, introduce un número.")

def procesar_datos(fecha, comunidad):
    archivo_destino = copiar_plantilla(fecha)
    wb = load_workbook(archivo_destino)
    ws = wb[comunidad]

    fecha_str = fecha.strftime('%Y-%m-%d')
    
    for row in range(2, ws.max_row + 1):
        try:
            coordenadas = ws[f'C{row}'].value
            if coordenadas:
                latitud, longitud = map(float, coordenadas.split(', '))
                datos = obtener_datos_meteo(latitud, longitud, fecha)

                if datos:
                    if fecha_str in datos['daily']['time']:
                        index = datos['daily']['time'].index(fecha_str)

                        ws[f'H{row}'] = datos['daily']['temperature_2m_min'][index]
                        ws[f'I{row}'] = datos['daily']['temperature_2m_max'][index]
                        ws[f'J{row}'] = datos['daily']['windspeed_10m_max'][index]
                        ws[f'K{row}'] = datos['daily']['windgusts_10m_max'][index]
                        ws[f'L{row}'] = datos['daily']['winddirection_10m_dominant'][index]
                        ws[f'M{row}'] = datos['daily']['precipitation_sum'][index]
                        ws[f'N{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        print(f"Fecha {fecha_str} no encontrada en los datos meteorológicos.")
        except Exception as e:
            print(f"Error al procesar la fila {row}: {e}")

    wb.save(archivo_destino)
    print(f"Datos meteorológicos guardados en {archivo_destino}")

def menu_principal():
    latitud, longitud = 37.3891, -5.9845
    comunidad = seleccionar_comunidad()
    
    fechas_disponibles = obtener_fechas_disponibles_api(latitud, longitud)
    if not fechas_disponibles:
        print("No se encontraron fechas disponibles.")
        return

    while True:
        print("\n--- Menú Principal ---")
        print("1. Seleccionar fecha y procesar datos")
        print("2. Opciones avanzadas de parametrización")
        print("3. Salir")
        opcion = input("Selecciona una opción: ")

        if opcion == '1':
            print("\nFechas disponibles:")
            for i, fecha in enumerate(fechas_disponibles, start=1):
                existe, ultima_actualizacion, datos_rellenos = verificar_archivo_existente(fecha, comunidad)
                estado = f"(Existente - Última actualización: {ultima_actualizacion}, Datos llenos: {datos_rellenos})" if existe else "(No existe)"
                print(f"{i}. {fecha} {estado}")

            while True:
                try:
                    seleccion_fecha = int(input("Selecciona la fecha (por número): ")) - 1
                    if 0 <= seleccion_fecha < len(fechas_disponibles):
                        fecha = datetime.strptime(fechas_disponibles[seleccion_fecha], '%Y-%m-%d')
                        break
                    else:
                        print("Selección no válida. Intenta nuevamente.")
                except ValueError:
                    print("Entrada no válida. Por favor, introduce un número.")

            procesar_datos(fecha, comunidad)

        elif opcion == '2':
            menu_opciones_avanzadas()

        elif opcion == '3':
            print("Saliendo del programa...")
            break

        else:
            print("Opción no válida. Por favor, elige nuevamente.")

# Ejecutar el menú principal
menu_principal()
