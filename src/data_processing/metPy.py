import os
import shutil
import requests
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import time
import json

def cls():
    """Limpia la pantalla."""
    os.system('cls' if os.name == 'nt' else 'clear')

def cargar_configuracion():
    """Carga la configuración desde el archivo config.json."""
    config_path = os.path.join(os.path.dirname(__file__), '..', '..', 'config', 'config.json')
    try:
        with open(config_path, "r", encoding="utf-8") as file:
            return json.load(file)
    except FileNotFoundError:
        print("Error: El archivo config.json no se encuentra en la ruta especificada.")
        input("Presione Enter para continuar...")
        exit(1)
    except json.JSONDecodeError:
        print("Error: El archivo config.json tiene un formato JSON no válido.")
        input("Presione Enter para continuar...")
        exit(1)

def cargar_traducciones(idioma="es"):
    """Carga el archivo de traducciones en el idioma especificado."""
    locales_dir = os.path.join(os.path.dirname(__file__), '..', '..', 'locales')
    idioma_path = os.path.join(locales_dir, f"{idioma}.json")
    try:
        with open(idioma_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print("Archivo de idioma no encontrado. Usando idioma predeterminado (español).")
        input("Presione Enter para continuar...")
        with open(os.path.join(locales_dir, 'es.json'), 'r', encoding='utf-8') as f:
            return json.load(f)

def inicializar_directorios(data_dir):
    """Inicializa el directorio de almacenamiento de datos."""
    os.makedirs(data_dir, exist_ok=True)

def obtener_fechas_disponibles_api(latitud, longitud):
    """Obtiene las fechas disponibles desde la API de Open-Meteo."""
    url = f"https://api.open-meteo.com/v1/forecast?latitude={latitud}&longitude={longitud}&daily=temperature_2m_max&timezone=Europe/Madrid"
    for intento in range(5):  # Máximo 5 reintentos
        response = requests.get(url)
        if response.status_code == 200:
            return response.json().get('daily', {}).get('time', [])
        elif response.status_code == 429:
            print(traducciones["too_many_requests"])
            time.sleep(5)
        else:
            print(traducciones["request_error"].format(code=response.status_code))
            input("Presione Enter para continuar...")
            return []
    print(traducciones["max_retries_reached"])
    input("Presione Enter para continuar...")
    return []

def verificar_archivo_existente(fecha, comunidad, data_dir):
    """Verifica si existe un archivo de datos para la fecha y comunidad especificadas."""
    fecha_str = fecha.replace("-", "")
    archivo = os.path.join(data_dir, f"{fecha[:4]}/{fecha[5:7]}/MeteoData_{fecha_str}.xlsx")
    if os.path.exists(archivo):
        ultima_actualizacion = datetime.fromtimestamp(os.path.getmtime(archivo)).strftime('%Y-%m-%d %H:%M:%S')
        datos_rellenos = verificar_datos_rellenos(archivo, comunidad)
        return True, ultima_actualizacion, datos_rellenos
    return False, None, False

def verificar_datos_rellenos(archivo, comunidad):
    """Verifica si hay datos rellenados en la hoja correspondiente a la comunidad autónoma."""
    wb = load_workbook(archivo)
    if comunidad not in wb.sheetnames:
        return False
    ws = wb[comunidad]
    return any(ws[f'H{row}'].value is not None for row in range(2, ws.max_row + 1))

def obtener_datos_meteo(latitud, longitud, fecha):
    """Obtiene los datos meteorológicos de la API para la fecha y coordenadas especificadas."""
    url = f"https://api.open-meteo.com/v1/forecast?latitude={latitud}&longitude={longitud}&daily=temperature_2m_max,temperature_2m_min,windspeed_10m_max,windgusts_10m_max,winddirection_10m_dominant,precipitation_sum&timezone=Europe/Madrid"
    response = requests.get(url)
    if response.status_code == 200:
        return response.json()
    else:
        print(traducciones["request_error"].format(code=response.status_code))
        input("Presione Enter para continuar...")
        return None

def copiar_plantilla(fecha, plantilla, data_dir):
    """Copia la plantilla de archivo Excel para la fecha especificada."""
    fecha_str = fecha.strftime('%Y%m%d')
    anio_dir = os.path.join(data_dir, fecha.strftime('%Y'))
    mes_dir = os.path.join(anio_dir, fecha.strftime('%m'))
    os.makedirs(mes_dir, exist_ok=True)
    archivo_destino = os.path.join(mes_dir, f"MeteoData_{fecha_str}.xlsx")
    shutil.copyfile(plantilla, archivo_destino)
    return archivo_destino

def seleccionar_comunidad(config):
    """Selecciona una comunidad autónoma de la configuración."""
    comunidades = config.get("comunidades", [])
    if not comunidades:
        print(traducciones["no_communities_found"])
        input("Presione Enter para continuar...")
        return None

    print(traducciones["select_community"])
    for i, comunidad in enumerate(comunidades, start=1):
        print(f"{i}. {comunidad}")

    while True:
        try:
            seleccion = int(input(traducciones["select_community_option"])) - 1
            if 0 <= seleccion < len(comunidades):
                return comunidades[seleccion]
            else:
                print(traducciones["invalid_selection"])
        except ValueError:
            print(traducciones["invalid_input"])

def procesar_datos(fecha, comunidad, data_dir, plantilla):
    """Procesa los datos meteorológicos para la fecha y comunidad especificadas."""
    archivo_destino = copiar_plantilla(fecha, plantilla, data_dir)
    wb = load_workbook(archivo_destino)
    ws = wb[comunidad]
    fecha_str = fecha.strftime('%Y-%m-%d')

    for row in range(2, ws.max_row + 1):
        try:
            coordenadas = ws[f'C{row}'].value
            if coordenadas:
                latitud, longitud = map(float, coordenadas.split(', '))
                datos = obtener_datos_meteo(latitud, longitud, fecha)

                if datos and fecha_str in datos['daily']['time']:
                    index = datos['daily']['time'].index(fecha_str)
                    ws[f'H{row}'] = datos['daily']['temperature_2m_min'][index]
                    ws[f'I{row}'] = datos['daily']['temperature_2m_max'][index]
                    ws[f'J{row}'] = datos['daily']['windspeed_10m_max'][index]
                    ws[f'K{row}'] = datos['daily']['windgusts_10m_max'][index]
                    ws[f'L{row}'] = datos['daily']['winddirection_10m_dominant'][index]
                    ws[f'M{row}'] = datos['daily']['precipitation_sum'][index]
                    ws[f'N{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                else:
                    print(traducciones["date_not_found"].format(date=fecha_str))
        except Exception as e:
            print(traducciones["processing_error"].format(row=row, error=str(e)))

    wb.save(archivo_destino)
    print(traducciones["data_saved"].format(archivo=archivo_destino))
    input("Presione Enter para continuar...")  # Pausa después de procesar datos

def menu_principal(config, plantilla):
    """Función principal del menú para procesar datos meteorológicos."""
    latitud, longitud = 37.3891, -5.9845
    comunidad = seleccionar_comunidad(config)

    fechas_disponibles = obtener_fechas_disponibles_api(latitud, longitud)
    if not fechas_disponibles:
        print(traducciones["no_dates_found"])
        input("Presione Enter para continuar...")
        return

    while True:
        cls()  # Limpia la pantalla
        print("\n--- " + traducciones["main_menu"] + " ---")
        print("1. " + traducciones["process_data"])
        print("2. " + traducciones["exit"])
        opcion = input(traducciones["select_option"])

        if opcion == '1':
            print("\n" + traducciones["available_dates"])
            for i, fecha in enumerate(fechas_disponibles, start=1):
                existe, ultima_actualizacion, datos_rellenos = verificar_archivo_existente(fecha, comunidad, data_dir)
                estado = (f"({traducciones['exists']} - {traducciones['last_update']}: {ultima_actualizacion}, "
                          f"{traducciones['data_filled']}: {datos_rellenos})" if existe
                          else f"({traducciones['does_not_exist']})")
                print(f"{i}. {fecha} {estado}")

            while True:
                try:
                    seleccion_fecha = int(input(traducciones["select_date_option"])) - 1
                    if 0 <= seleccion_fecha < len(fechas_disponibles):
                        fecha = datetime.strptime(fechas_disponibles[seleccion_fecha], '%Y-%m-%d')
                        break
                    else:
                        print(traducciones["invalid_selection"])
                except ValueError:
                    print(traducciones["invalid_input"])

            procesar_datos(fecha, comunidad, data_dir, plantilla)

        elif opcion == '2':
            print(traducciones["exiting"])
            break
        else:
            print(traducciones["invalid_selection"])
            input("Presione Enter para continuar...")  # Pausa si la selección no es válida

# Cargar la configuración y traducciones
config = cargar_configuracion()
traducciones = cargar_traducciones(config.get("language", "es"))

# Directorio de almacenamiento de datos
data_dir = config.get("data_dir", "data")
inicializar_directorios(data_dir)

# Parámetros avanzados desde el archivo de configuración
plantilla = config.get("template_path", "config/templates/MeteoData.xlsx")

# Ejecutar el menú principal
menu_principal(config, plantilla)
